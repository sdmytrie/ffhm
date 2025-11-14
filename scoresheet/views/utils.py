from collections import OrderedDict
from copy import deepcopy
import math
from operator import attrgetter

from django.db.models import Avg, FloatField, Max, Min, Q
from django.db.models.functions import Cast
from icecream import ic
from openpyxl import Workbook
from openpyxl.styles import Color, Font, PatternFill, colors
from openpyxl.writer.excel import save_virtual_workbook

from api.models import *
import pymongo


def provisional_iwf(event):
    iwf = 0
    total = 0
    epj_value = 0
    arr_value = 0
    for attempt in event.attempt_set.all():
        if attempt.validate == 2:
            continue
        if attempt.name == "EP-J":
            if attempt.value > epj_value:
                epj_value = attempt.value
        else:
            if attempt.value > arr_value:
                arr_value = attempt.value
    total = epj_value + arr_value
    if event.concurrent.gender_id == 2:
        # A = 0.751945030
        # b = 175.508
        A = 0.722762521
        b = 193.609
    else:
        # A = 0.783497476
        # b = 153.655
        A = 0.787004341
        b = 153.757

    if (
        event.competition.isteam
        and event.competition.isminime
        and event.competition.kind.id == 12
    ):
        A = 0.722762521
        b = 193.609

    X = float(event.weight) / b
    if X > 0:
        iwf = pow(10, A * pow(math.log10(X), 2)) * total

    return iwf


def iwf(event):
    iwf = 0

    total = event.totalSet[0] + event.totalSet[1]
    if not event.competition.isteam:
        if event.totalSet[0] == 0 or event.totalSet[1] == 0:
            total = 0

    def truncate(number, digits) -> float:
        stepper = pow(10.0, digits)
        return math.trunc(stepper * number) / stepper

    if event.totalSet[1] > 0 or event.competition.isteam:
        if event.concurrent.gender_id == 2:
            # A = 0.751945030
            # b = 175.508
            A = 0.722762521
            b = 193.609
        else:
            # A = 0.783497476
            # b = 153.655
            A = 0.787004341
            b = 153.757

        if (
            event.competition.isteam
            and event.competition.isminime
            and event.competition.kind.id == 12
        ):
            A = 0.722762521
            b = 193.609

        X = float(event.weight) / b
        if X > 0:
            iwf = pow(10, A * pow(math.log10(X), 2)) * total

    return truncate(iwf, 2)


def sortClosedEventListOld(competition):
    u10EventList = Event.objects.filter(
        competition_id=competition.id, agecategory__name="U10"
    ).order_by("-concurrent__gender", "agecategory__name", "-total")
    # uEventList = Event.objects.annotate(myweight=Cast('weightcategory__weight', FloatField())).filter(competition_id = competition.id).order_by('agecategory__name', 'myweight', '-total').exclude(agecategory__name='SENIOR').exclude(agecategory__name='U10')

    uEventList = (
        Event.objects.filter(competition_id=competition.id)
        .order_by(
            "-concurrent__gender", "agecategory__name", "weightcategory__id", "-total"
        )
        .exclude(agecategory__name="SENIOR")
        .exclude(agecategory__name="U10")
    )

    seniorEventList = (
        Event.objects.annotate(myweight=Cast("weightcategory__id", FloatField()))
        .filter(competition_id=competition.id, agecategory__name="SENIOR")
        .order_by("-concurrent__gender", "myweight", "-total")
    )
    dictEvent = OrderedDict()

    for event in u10EventList:
        if event.agecategory.name not in dictEvent:
            dictEvent[event.agecategory.name] = list()
        dictEvent[event.agecategory.name].append(event)

    for event in uEventList:
        if event.agecategory.name not in dictEvent:
            dictEvent[event.agecategory.name] = list()
        dictEvent[event.agecategory.name].append(event)

    for event in seniorEventList:
        if event.agecategory.name not in dictEvent:
            dictEvent[event.agecategory.name] = list()
        dictEvent[event.agecategory.name].append(event)

    return dictEvent


def sort_closed_event_list(competition):
    """sort closed events"""

    buffer_event_list = OrderedDict()
    event_dict = OrderedDict()
    buffer_concurrent = dict()

    buffer_event_list = Event.objects.filter(competition_id=competition.id).order_by(
        "agecategory_id", "weightcategory_id", "-total"
    )

    event_list = sorted(
        buffer_event_list,
        # key=lambda e: (
        #     e.agecategory_id,
        #     e.weightcategory_id,
        #     -e.total,
        #     e.totalSet[1],
        #     e.iwf,
        # ),
        key=lambda e: (
            e.agecategory_id,
            e.weightcategory_id,
            -e.iwf,
        ),
    )

    for event in event_list:
        if event.weightcategory not in buffer_concurrent:
            buffer_concurrent[event.weightcategory] = OrderedDict()

        if event.concurrent not in buffer_concurrent[event.weightcategory]:
            buffer_concurrent[event.weightcategory][event.concurrent] = 0

        if event.concurrent.gender not in event_dict:
            event_dict[event.concurrent.gender] = OrderedDict()

        if event.agecategory not in event_dict[event.concurrent.gender]:
            event_dict[event.concurrent.gender][event.agecategory] = OrderedDict()

        if (
            event.weightcategory
            not in event_dict[event.concurrent.gender][event.agecategory]
        ):
            event_dict[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ] = OrderedDict()

        if (
            event.concurrent
            not in event_dict[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ]
        ):
            event_dict[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ][event.concurrent] = None

        if event.total >= buffer_concurrent[event.weightcategory][event.concurrent]:
            event_dict[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ][event.concurrent] = event
            buffer_concurrent[event.weightcategory][event.concurrent] = event.total

    return event_dict


# def sortEventList(competition):
#     bufferEventSet = OrderedDict()
#     eventSet = OrderedDict()
#     bufferConcurrent = dict()

#     bufferEventSet = list(Event.objects.filter(competition_id=competition.id).order_by('draw', 'agecategory_id', 'weightcategory_id'))

#     # factory = lambda: defaultdict(factory)
#     # eventList = factory()

#     for event in bufferEventSet:
#         if event.weightcategory not in bufferConcurrent:
#             bufferConcurrent[event.weightcategory] = OrderedDict()

#         if event.concurrent not in bufferConcurrent[event.weightcategory]:
#             bufferConcurrent[event.weightcategory][event.concurrent] = 0

#         if event.concurrent.gender not in eventSet:
#             eventSet[event.concurrent.gender] = OrderedDict()

#         if event.agecategory not in eventSet[event.concurrent.gender]:
#             eventSet[event.concurrent.gender][event.agecategory] = OrderedDict()

#         if event.weightcategory not in eventSet[event.concurrent.gender][event.agecategory]:
#             eventSet[event.concurrent.gender][event.agecategory][event.weightcategory] = OrderedDict()

#         if  event.concurrent not in eventSet[event.concurrent.gender][event.agecategory][event.weightcategory]:
#             eventSet[event.concurrent.gender][event.agecategory][event.weightcategory][event.concurrent] = None

#         # if event.total >= bufferConcurrent[event.weightcategory][event.concurrent]:
#         eventSet[event.concurrent.gender][event.agecategory][event.weightcategory][event.concurrent] = event
#             # bufferConcurrent[event.weightcategory][event.concurrent] = event.total

#     return eventSet

# def sortEventList(eventList):


def sortEventListOld(eventList):
    # eventList = list(competition.event_set.all())
    attemptList = []
    garbageList = []

    for event in eventList:
        attemptList = attemptList + list(
            event.attempt_set.filter(validate=0, value__gt=0, name="ARR")
        )
        garbageList = garbageList + list(event.attempt_set.filter(validate=2))
        garbageList = garbageList + list(event.attempt_set.filter(validate=1))

    if not attemptList:
        for event in eventList:
            attemptList = attemptList + list(
                event.attempt_set.filter(validate=0, value__gt=0, name="EP-J")
            )

    if not attemptList:
        # eventList.sort(key=attrgetter('concurrent.date_of_birth'), reverse = True)
        # bufferList = sorted(eventList, key = attrgetter('concurrent.date_of_birth.year'))

        # bufferList = sorted(eventList, key = lambda e: e.concurrent.date_of_birth.year)
        eventList = sorted(eventList, key=attrgetter("iwf"), reverse=True)
    else:
        myList = sorted(
            attemptList, key=attrgetter("value", "rank", "distance", "event.draw")
        )
        gbList = sorted(
            garbageList, key=attrgetter("value", "distance", "event.draw", "rank")
        )

        for event in eventList:
            myList = myList + list(
                event.attempt_set.filter(validate=0, value=0, rank=1)
            )

        eventList = []
        for attempt in myList:
            if attempt.event in eventList:
                continue
            eventList.append(attempt.event)

        for attempt in gbList:
            if attempt.event in eventList:
                continue
            eventList.append(attempt.event)
    return eventList


def sortEventTeamList(eventList):
    # eventList = list(self.event_set.all())
    attemptList = []
    garbageList = []

    femaleAttemptCount = 0
    for event in eventList:
        if event.concurrent.gender_id == 2:
            continue
        femaleAttemptCount += event.attempt_set.filter(validate=0, value__gt=0).count()

    for event in eventList:
        if femaleAttemptCount > 0 and event.concurrent.gender.id == 2:
            continue
        attemptList = attemptList + list(
            event.attempt_set.filter(validate=0, value__gt=0, name="ARR")
        )
        garbageList = garbageList + list(event.attempt_set.filter(validate=2))
        garbageList = garbageList + list(event.attempt_set.filter(validate=1))

    if not attemptList:
        for event in eventList:
            attemptList = attemptList + list(
                event.attempt_set.filter(validate=0, value__gt=0, name="EP-J")
            )

    if not attemptList:
        return sorted(eventList, key=lambda e: e.iwf or e.total, reverse=True)

    myList = sorted(
        attemptList,
        key=attrgetter("value", "rank", "distance", "event.team.draw", "event.id"),
    )
    gbList = sorted(
        garbageList,
        key=attrgetter("value", "distance", "event.draw", "rank", "event.team.draw"),
    )

    for event in eventList:
        myList = myList + list(event.attempt_set.filter(validate=0, value=0, rank=1))

    eventList = []
    for attempt in myList:
        if attempt.event in eventList:
            continue
        eventList.append(attempt.event)

    for attempt in gbList:
        if attempt.event in eventList:
            continue
        eventList.append(attempt.event)

    return eventList


#
# PERMISSIONS
#
def addAuthorization(model, authorization, user):
    if user.is_authenticated:
        content_type = ContentType.objects.get(app_label="scoresheet", model=model)
        permission = Permission.objects.get(
            codename=authorization, content_type_id=content_type.id
        )
        user.user_permissions.add(permission)


def removeAuthorization(model, authorization, user):
    if user.is_authenticated:
        content_type = ContentType.objects.get(app_label="scoresheet", model=model)
        permission = Permission.objects.get(
            codename=authorization, content_type_id=content_type.id
        )
        user.user_permissions.remove(permission)


def isEditor(user):
    return user.groups.filter(name="Editor").exists()


#
# Excel
#
def excelize(event_list, title):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title

    columns = [
        "Licence",
        "Nom",
        "AN",
        "Club",
        "NAT",
        "P.C.",
        "1",
        "2",
        "3",
        "ARR",
        "1",
        "2",
        "3",
        "EP-J",
        "TOTAL",
        "Série",
        "Catégorie",
        "IWF",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid"
        )
        cell.font = Font(color=colors.WHITE)

    for gender, agecategory_list in event_list.items():
        row_num += 1
        row = [gender.verbosename.upper()]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
        for agecategory, weightcategory_list in agecategory_list.items():
            row_num += 1
            row = [agecategory.name]
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
            for weightcategory, concurrent_list in weightcategory_list.items():
                for concurrent, event in concurrent_list.items():
                    row_num += 1

                    # worksheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=18)
                    # # Define the data for each cell in the row
                    row = [
                        event,
                    ]

                    # # Assign the data for each cell of the row
                    # for col_num, cell_value in enumerate(row, 1):
                    #     cell = worksheet.cell(row=row_num, column=col_num)
                    #     cell.value = cell_value

                    # # worksheet.unmerge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=18)
                    # for event in value:
                    row = [
                        event.concurrent.licence,
                        event.concurrent.lastname.upper()
                        + " "
                        + event.concurrent.firstname.capitalize(),
                        event.concurrent.date_of_birth.year,
                        event.concurrent.clubName,
                        event.concurrent.country,
                        event.weight,
                    ]
                    for attempt in event.getARRAttemptSet:
                        row.append(((-1) ** (attempt.validate + 1)) * attempt.value)
                        cell = worksheet.cell(row=row_num, column=6 + attempt.rank)
                        color = "00FF00"
                        if attempt.validate == 2:
                            color = "FF0000"
                        else:
                            color = "00FF00"
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                    row.append(event.totalSet[0])

                    for attempt in event.getEPJAttemptSet:
                        row.append(((-1) ** (attempt.validate + 1)) * attempt.value)
                        cell = worksheet.cell(row=row_num, column=10 + attempt.rank)
                        color = "00FF00"
                        if attempt.validate == 2:
                            color = "FF0000"
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )
                    row.append(event.totalSet[1])
                    row.append(event.total)

                    if event.minimumweightcategory and not event.competition.isminime:
                        row.append(event.minimumweightcategory.name)
                    else:
                        row.append("N.C.")

                    if event.competition.ismasters:
                        row.append(
                            event.agecategory
                            + " "
                            + event.categorie
                            + " "
                            + event.weightcategory.weight
                        )
                    else:
                        if (
                            "W" in event.agecategory.name
                            or "M" in event.agecategory.name
                        ):
                            categorie = event.concurrent.gender.name[:1].upper()
                            row.append(
                                "MASTERS "
                                + categorie
                                + " "
                                + event.weightcategory.weight
                            )
                        else:
                            categorie = event.concurrent.gender.name[:1].upper()
                            if event.weightcategory:
                                row.append(
                                    event.agecategory.name[:3]
                                    + " "
                                    + categorie
                                    + " "
                                    + event.weightcategory.weight
                                )
                            else:
                                row.append(
                                    event.agecategory.name[:3] + " " + categorie + " "
                                )

                    row.append(event.iwf)

                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = (
                length * 1.5
            )

    return workbook


def get_clubs_from_region(region):
    collection = pymongo.MongoClient("mongo", 27017).exalto.concurrent
    concurrents = list(collection.find({"concurrent.result.club.region.nom": region}))
    clubs = []
    for concurrent in concurrents:
        if concurrent["concurrent"]["result"]["club"]["nom"] not in clubs:
            clubs.append(concurrent["concurrent"]["result"]["club"]["nom"])

    return clubs


def get_region_by_concurrent(concurrent):
    collection = pymongo.MongoClient("mongo", 27017).exalto.concurrent
    current = collection.find_one(
        {"concurrent.result.code_adherent": concurrent.licence}
    )

    if current:
        return (
            current.get("concurrent", {})
            .get("result", {})
            .get("club", {})
            .get("region", {})
            .get("nom", "")
        )
    return ""


class ManageRecords:
    def get_events(self):
        event_list = list(
            Event.objects.filter(
                competition__isrecordeligible=True,
                competition__closed=True,
                competition__isminime=False,
                concurrent__country="FR",
            ).all()
        )
        event_list.sort(key=lambda x: x.updated_at)

        return event_list

    def set_weightcategory(self, event: Event) -> Event:
        current_weightcategoryList = []
        current_weightcategory = None
        max_weightcategory = None
        current_season = list(Season.objects.order_by("-id").all())

        if event.competition.season.pk == current_season[0].pk:
            return event

        current_agecategory = Agecategory.objects.get(
            name=event.agecategory.name,
            season=current_season[0],
            gender=event.competition.gender,
        )
        if float(event.weight) == 0.0 or event.agecategory.name == "U10":
            current_weightcategory = None
        else:
            weightcategoryList = list(
                Weightcategory.objects.filter(agecategory_id=current_agecategory.pk)
            )
            for weightcategory in weightcategoryList:
                if ">" in weightcategory.weight:
                    max_weightcategory = weightcategory
                    continue
                current_weightcategoryList.append(weightcategory)

            current_weightcategoryList.sort(key=lambda w: float(w.weight))
            current = weightcategoryList[0]

            for weightcategory in current_weightcategoryList:
                if float(event.weight) <= float(current.weight):
                    current_weightcategory = current
                    break
                current = weightcategory

            current_max_weightcategory = current_weightcategoryList[-1]
            if not current_weightcategory:
                current_weightcategory = current_max_weightcategory

            if float(event.weight) > float(current_max_weightcategory.weight):
                current_weightcategory = max_weightcategory
        event.weightcategory = current_weightcategory

        return event

    def get_last_agecategory(self, event: Event) -> Event | None:
        age = (
            event.competition.season.end_date.year - event.concurrent.date_of_birth.year
        )
        current_season = list(Season.objects.order_by("-id").all())

        agecategoryList = list(
            Agecategory.objects.filter(
                season_id=event.competition.season.id,
                gender_id=event.concurrent.gender_id,
            )
            .exclude(name__startswith="M")
            .exclude(name__startswith="W")
            .order_by("agemin")
        )

        current = agecategoryList[0]
        new_event = None
        existing_agecategory = ["SENIOR", "U20", "U17", "U15"]
        for agecategory in agecategoryList:
            if age - 1 <= current.agemax:
                if (
                    event.agecategory.name != current.name
                    and event.updated_at.year == current_season[0].start_date.year
                    and current.name in existing_agecategory
                ):
                    new_event = deepcopy(event)
                    new_event.agecategory = current
                break
            current = agecategory
        return new_event
