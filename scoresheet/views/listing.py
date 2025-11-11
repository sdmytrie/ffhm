"""
LISTING
"""

from collections import OrderedDict, defaultdict
import copy
from datetime import date, datetime

from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from openpyxl import Workbook

from api.models import (
    Agecategory,
    Competition,
    Event,
    Gender,
    Leader,
    Minimumweightcategory,
    Season,
    Weightcategory,
)
from scoresheet.views import utils
from scoresheet.forms import ListingForm
from scoresheet.views.utils import get_region_by_concurrent


@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def listing(request, season_id="0"):
    """listing"""

    title = "Listing"
    page = "listing"

    current_season = list(
        Season.objects.all().order_by("start_date").reverse())[0]
    if season_id == "0":
        season = current_season
    else:
        season = Season.objects.get(id=season_id)

    season_url_value = reverse("scoresheet:listing")
    season_list = Season.objects.all().order_by("start_date")

    series = [
        "N.C",
        "DEB",
        "DPT",
        "REG",
        "IRG",
        "FED",
        "HON",
        "NAT",
        "INT B",
        "INT A",
        "EUR",
        "MON",
        "OLY",
        "MONDE",
        "EUROPE",
        "FRANCE",
        "REGION",
    ]
    ages = ["U10", "U13", "U15", "U17", "U20", "SENIOR", "Masters"]

    if request.POST.get("startDate"):
        start_date = datetime.strptime(
            request.POST["startDate"], "%d/%m/%Y").date()
    else:
        start_date = season.start_date

    if request.POST.get("endDate"):
        end_date = datetime.strptime(
            request.POST["endDate"], "%d/%m/%Y").date()
    else:
        end_date = season.end_date

    if request.user.is_superuser:
        competitions = list(
            Competition.objects.prefetch_related("gender", "user")
            .filter(
                # season_id=season.id,
                closed=True,
                start_date__range=(start_date, end_date),
            )
            .order_by("-start_date")
        )
    elif request.user.is_staff:
        competitions = list(
            Competition.objects.prefetch_related("gender", "user")
            .filter(
                # season_id=season.id,
                closed=True,
                start_date__range=(start_date, end_date),
                user__profile__region=request.user.profile.region,
            )
            .order_by("-start_date")
        )

        clubs = utils.get_clubs_from_region(request.user.profile.region)
        events = list(
            Event.objects.filter(
                concurrent__clubName__in=clubs,
                competition__closed=True,
                competition__start_date__range=(start_date, end_date),
            )
        )
        for event in events:
            buffer_competition = Competition.objects.prefetch_related(
                "gender", "user", "gender"
            ).get(id=event.competition.id)
            if buffer_competition not in competitions:
                competitions.append(buffer_competition)

    date_list = []
    for competition in competitions:
        if competition.start_date not in date_list:
            date_list.append(competition.start_date)

    return render(request, "scoresheet/listing/listing.html", locals())


@user_passes_test(lambda u: u.is_superuser or u.is_staff)
def listing_view(request):
    """Description"""

    title = "Listing"
    page = "listing"
    current_season = list(
        Season.objects.all().order_by("start_date").reverse())[0]

    buffer_event_set = OrderedDict()
    event_set = OrderedDict()
    buffer_concurrent = dict()

    gender_list = list()
    gender_list = request.POST.get("genderList").split("|")

    age_list = list()
    my_age_list = list()
    my_age_list = request.POST.get("ageList").split("|")
    master_list = [
        "W35",
        "W40",
        "W45",
        "W50",
        "W55",
        "W60",
        "W65",
        "W70",
        "W75",
        "W80",
        "W85",
        "M35",
        "M40",
        "M45",
        "M50",
        "M55",
        "M60",
        "M65",
        "M70",
        "M75",
        "M80",
        "M85",
    ]

    is_senior = False
    if "SENIOR" in my_age_list:
        is_senior = True

    is_master = False
    if "Masters" in my_age_list:
        my_age_list = my_age_list + master_list
        is_master = True

    age_list = Agecategory.objects.filter(name__in=my_age_list)

    my_serie_list = request.POST.get("serieList").split("|")
    serie_list = list()
    serie_null = False
    if "N.C" in my_serie_list:
        serie_null = True
    serie_list = Minimumweightcategory.objects.filter(name__in=my_serie_list)

    clubs = "all"
    if not request.user.is_superuser:
        clubs = utils.get_clubs_from_region(request.user.profile.region)

    query = Q(competition_id__in=request.POST.get("competitionSet").split(","))
    # query.add(Q(competition__season_id=current_season.id), Q.AND)
    query.add(Q(concurrent__gender_id__in=gender_list), Q.AND)
    if not request.user.is_superuser:
        query.add(Q(concurrent__clubName__in=clubs), Q.AND)
    query.add(Q(agecategory_id__in=age_list), Q.AND)

    if serie_null:
        query.add(
            Q(minimumweightcategory_id__in=serie_list)
            | Q(minimumweightcategory__isnull=True),
            Q.AND,
        )
    else:
        query.add(Q(minimumweightcategory_id__in=serie_list), Q.AND)
    buffer_event_set = list(Event.objects.filter(query))

    if is_master:
        age_list = Agecategory.objects.filter(name="SENIOR")
        query = Q(competition_id__in=request.POST.get(
            "competitionSet").split(","))
        # query.add(Q(competition__season_id=current_season.id), Q.AND)
        query.add(Q(concurrent__gender_id__in=gender_list), Q.AND)
        query.add(Q(agecategory_id__in=age_list), Q.AND)

        master_event_set = list(
            Event.objects.prefetch_related("competition").filter(query)
        )

        master_age_list = [
            "35",
            "40",
            "45",
            "50",
            "55",
            "60",
            "65",
            "70",
            "75",
            "80",
            "85",
        ]
        for event in master_event_set:
            # nb_years = current_season.start_date.year - event.concurrent.date_of_birth.year
            nb_years = (
                event.competition.season.start_date.year
                - event.concurrent.date_of_birth.year
            )
            for master_age in master_age_list:
                min = int(master_age)
                if (nb_years + 1 >= min and nb_years + 1 < (min + 5)) or (
                    min == 85 and nb_years + 1 > 90
                ):
                    if event.concurrent.gender_id == 2:
                        agecategory_name = "M" + master_age
                    else:
                        agecategory_name = "W" + master_age
                        if int(master_age) > 85:
                            agecategory_name = "W85"
                    # agecategory_name = 'M35'
                    new_event = copy.deepcopy(event)
                    b_age = Agecategory.objects.get(
                        season_id=event.competition.season.id,
                        gender_id=event.agecategory.gender.id,
                        name=agecategory_name,
                    )
                    new_event.agecategory = b_age
                    b_min_set = Minimumweightcategory.objects.filter(
                        weightcategory__agecategory__season_id=event.competition.season.id,
                        weightcategory__agecategory__gender_id=event.agecategory.gender.id,
                        weightcategory__agecategory_id=b_age.id,
                        weightcategory__weight=event.weightcategory.weight,
                        weightcategory__agecategory__name=agecategory_name,
                    ).order_by("weight")

                    b_weightcategory = Weightcategory.objects.get(
                        agecategory__season_id=event.competition.season.id,
                        agecategory__gender_id=event.agecategory.gender.id,
                        weight=event.weightcategory.weight,
                        agecategory__name=agecategory_name,
                    )
                    new_event.weightcategory = b_weightcategory

                    current = None
                    new_event.minimumweightcategory = None
                    for b_min in b_min_set:
                        if new_event.total >= b_min.weight:
                            new_event.minimumweightcategory = b_min

                    if new_event.minimumweightcategory is not None:
                        if new_event.minimumweightcategory.name in my_serie_list:
                            buffer_event_set.append(new_event)
                    else:
                        if serie_null:
                            buffer_event_set.append(new_event)

    age_list = list()
    if is_senior:
        age_list = Agecategory.objects.filter(name__in=master_list)
        query = Q(competition_id__in=request.POST.get(
            "competitionSet").split(","))
        # query.add(Q(competition__season_id=current_season.id), Q.AND)
        query.add(Q(concurrent__gender_id__in=gender_list), Q.AND)
        query.add(Q(agecategory_id__in=age_list), Q.AND)

        master_event_set = list(Event.objects.filter(query))
        for event in master_event_set:
            new_event = copy.deepcopy(event)

            b_age = Agecategory.objects.get(
                season_id=event.competition.season.id,
                gender_id=event.agecategory.gender.id,
                name="SENIOR",
            )
            new_event.agecategory = b_age

            b_min_set = Minimumweightcategory.objects.filter(
                weightcategory__agecategory__season_id=event.competition.season.id,
                weightcategory__agecategory__gender_id=event.agecategory.gender.id,
                weightcategory__agecategory_id=b_age.id,
                weightcategory__weight=event.weightcategory.weight,
                weightcategory__agecategory__name="SENIOR",
            ).order_by("weight")

            b_weightcategory = Weightcategory.objects.get(
                agecategory__season_id=event.competition.season.id,
                agecategory__gender_id=event.agecategory.gender.id,
                weight=event.weightcategory.weight,
                agecategory__name="SENIOR",
            )
            new_event.weightcategory = b_weightcategory

            current = None
            for b_min in b_min_set:
                if new_event.total >= b_min.weight:
                    new_event.minimumweightcategory = b_min
                # current = b_min

            if new_event.minimumweightcategory is not None:
                if new_event.minimumweightcategory.name in my_serie_list:
                    buffer_event_set.append(new_event)
            else:
                if serie_null:
                    buffer_event_set.append(new_event)

    event_list = sorted(
        buffer_event_set,
        key=lambda e: (e.agecategory_id,
                       e.weightcategory_id,
                       -e.total, e.totalSet[1]),
    )

    for event in event_list:
        event.league = get_region_by_concurrent(event.concurrent)
        if event.weightcategory not in buffer_concurrent:
            buffer_concurrent[event.weightcategory] = OrderedDict()

        if event.concurrent not in buffer_concurrent[event.weightcategory]:
            buffer_concurrent[event.weightcategory][event.concurrent] = 0
        # else:
        #     buffer_concurrent[event.weightcategory][event.concurrent] = event.total

        if event.concurrent.gender not in event_set:
            event_set[event.concurrent.gender] = OrderedDict()

        agecategory_name = event.agecategory.name
        if event.agecategory.name.startswith("W") or event.agecategory.name.startswith(
            "M"
        ):
            agecategory_name = "Masters"

        if agecategory_name not in event_set[event.concurrent.gender]:
            event_set[event.concurrent.gender][agecategory_name] = OrderedDict()

        if (
            event.weightcategory
            not in event_set[event.concurrent.gender][agecategory_name]
        ):
            event_set[event.concurrent.gender][agecategory_name][
                event.weightcategory
            ] = OrderedDict()

        if (
            event.concurrent
            not in event_set[event.concurrent.gender][agecategory_name][
                event.weightcategory
            ]
        ):
            event_set[event.concurrent.gender][agecategory_name][event.weightcategory][
                event.concurrent
            ] = None

        if event.total >= buffer_concurrent[event.weightcategory][event.concurrent]:
            event_set[event.concurrent.gender][agecategory_name][event.weightcategory][
                event.concurrent
            ] = event
            buffer_concurrent[event.weightcategory][event.concurrent] = event.total

    statistics = dict()
    concurrents = []
    for gender, agecategory_set in event_set.items():
        for agecategory_name, weightcategory_set in agecategory_set.items():
            statistics[agecategory_name] = 0
            for weightcategory, concurrent_set in weightcategory_set.items():
                statistics[agecategory_name] += len(concurrent_set)
                for concurrent, event in concurrent_set.items():
                    if concurrent not in concurrents:
                        concurrents.append(concurrent)
                    else:
                        statistics[agecategory_name] -= 1
                        event.duplicate = True

    return render(request, "scoresheet/listing/listing_view.html", locals())


def default_to_regular(d):
    if isinstance(d, defaultdict):
        d = {k: default_to_regular(v) for k, v in d.items()}
    return d


@user_passes_test(lambda u: u.is_superuser)
def listing_form(request):
    """Description"""

    title = "Listing"
    # competition = Competition.objects.get(id=id_competition)
    # isteam = competition.isteam

    if request.method == "POST":
        form = ListingForm(request.POST)
        # if form.is_valid():
        return HttpResponseRedirect("/scoresheet/listing")
    else:
        form = ListingForm()

    return render(request, "scoresheet/ListingForm.html", locals())


@user_passes_test(lambda u: u.is_superuser)
def listingXls(request, id_competition):
    """Description"""

    bufferEventSet = OrderedDict()
    eventSet = OrderedDict()
    bufferConcurrent = dict()

    bufferEventSet = list(
        Event.objects.filter(
            competition_id__in=request.POST.get("competitionSet").split(",")
        ).order_by("agecategory_id", "weightcategory_id", "-total")
    )

    # factory = lambda: defaultdict(factory)
    # eventList = factory()

    for event in bufferEventSet:
        if event.concurrent not in bufferConcurrent:
            bufferConcurrent[event.weightcategory][event.concurrent] = 0

        if event.concurrent.gender not in eventSet:
            eventSet[event.concurrent.gender] = OrderedDict()

        if event.agecategory not in eventSet[event.concurrent.gender]:
            eventSet[event.concurrent.gender][event.agecategory] = OrderedDict()

        if (
            event.weightcategory
            not in eventSet[event.concurrent.gender][event.agecategory]
        ):
            eventSet[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ] = OrderedDict()

        if (
            event.concurrent
            not in eventSet[event.concurrent.gender][event.agecategory][
                event.weightcategory
            ]
        ):
            eventSet[event.concurrent.gender][event.agecategory][event.weightcategory][
                event.concurrent
            ] = None

        if event.total > bufferConcurrent[event.weightcategory][event.concurrent]:
            eventSet[event.concurrent.gender][event.agecategory][event.weightcategory][
                event.concurrent
            ] = event
            bufferConcurrent[event.weightcategory][event.concurrent] = event.total

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = competition.name

    columns = [
        "Licence",
        "Nom",
        "AN",
        "Club",
        "NAT",
        "P.C.",
        "1",
        "2",
        "3",
        "ARR",
        "1",
        "2",
        "3",
        "EP-J",
        "TOTAL",
        "Série",
        "Catégorie",
        "IWF",
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for key, value in eventSet.items():
        row_num += 1

        worksheet.merge_cells(
            start_row=row_num, start_column=1, end_row=row_num, end_column=18
        )
        # Define the data for each cell in the row
        row = [
            key,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

        # worksheet.unmerge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=18)
        for event in value:
            row_num += 1
            row = [
                event.concurrent.licence,
                event.concurrent.lastname.upper()
                + " "
                + event.concurrent.firstname.capitalize(),
                event.concurrent.date_of_birth.year,
                event.concurrent.clubName,
                event.concurrent.country,
                event.weight,
            ]
            for attempt in event.getARRAttemptSet:
                row.append(attempt.value)
            row.append(event.totalSet[0])

            for attempt in event.getEPJAttemptSet:
                row.append(attempt.value)
            row.append(event.totalSet[1])
            row.append(event.total)

            if event.minimumweightcategory and not competition.isminime:
                row.append(event.minimumweightcategory.name)
            else:
                row.append("N.C.")

            if event.competition.ismasters:
                row.append(
                    event.agecategory
                    + " "
                    + event.categorie
                    + " "
                    + event.weightcategory.weight
                )
            else:
                if "W" in event.agecategory.name or "M" in event.agecategory.name:
                    categorie = event.concurrent.gender.name[:1].upper()
                    row.append(
                        "MASTERS " + categorie + " " + event.weightcategory.weight
                    )
                else:
                    categorie = event.concurrent.gender.name[:1].upper()
                    if event.weightcategory:
                        row.append(
                            event.agecategory.name[:3]
                            + " "
                            + categorie
                            + " "
                            + event.weightcategory.weight
                        )
                    else:
                        row.append(
                            event.agecategory.name[:3] + " " + categorie + " ")

            row.append(event.iwf)

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


def listing_leader(request, season_id="0"):
    """list all leaders"""

    title = "Officiels"
    page = "listing_leader"

    current_season = list(
        Season.objects.all().order_by("start_date").reverse())[0]
    if season_id == "0":
        season = current_season
    else:
        season = Season.objects.get(id=season_id)
    season_url_value = reverse("scoresheet:listing_leader")

    season_list = Season.objects.all().order_by("start_date")
    seasonUrlValue = ""

    leaders = (
        Leader.objects.prefetch_related("competition")
        .filter(competition__season__id=season.id)
        .order_by("concurrent__lastname", "competition__start_date")
    )
    # leader_dict = dict()

    # for leader in leaders:
    #     buffer_list = (leader.concurrent, leader.competition)
    #     if leader.concurrent not in leader_dict:
    #         leader_dict[leader.concurrent] = list()
    # leader_dict[leader.concurrent].append(buffer_list)

    return render(request, "scoresheet/listing/listing_leader.html", locals())
